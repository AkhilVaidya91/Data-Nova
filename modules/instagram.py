####################### IMPORTING REQIRED LIBRARIES #######################
import numpy as np
import pandas as pd
import requests
import os
import tempfile
import google.generativeai as genai
from apify_client import ApifyClient
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

####################### FUNCTION DEFINITIONS #######################

def remove_comma(text):
    text =  text.replace(",", " ")
    return str(text)

def setup(api_key):
    client = ApifyClient(api_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Instagram Posts"
    headers = ["Platform", "Caption", "Hashtags", "Alt Text", "Post Type", "Post/Image URL", "Likes", "Comment count", "Day", "Month", "Year", "Company Name", "Is Sponsored", "Generated Caption"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return client, ws, wb

def setup_profile(api_key):
    client = ApifyClient(api_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Instagram Profile"
    headers = ["User Name", "Full Name", "Biography", "Subscriber Count", "Following Count", "Reels Count - video", "Posts Count - image"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return client, ws, wb

def num_months_to_posts(n):
    if n < 7:
        return 70
    elif n < 12:
        return 120
    else:
        return 500


def calculate_previous_date(day, month, year, months_delta):
    # Create the initial date
    initial_date = datetime(year, month, day)
    
    # Calculate the new date by subtracting months_delta months
    new_date = initial_date - relativedelta(months=months_delta)
    
    # Return the new date as (year, month, day)
    return new_date.date()

def get_post_text(post_url, api_key):
    genai.configure(api_key=api_key)
    response = requests.get(post_url)
    image_data = response.content

    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_file:
        temp_file.write(image_data)
        temp_file_path = temp_file.name

    try:
        myfile = genai.upload_file(temp_file_path)
        model = genai.GenerativeModel("gemini-1.5-flash")
        result = model.generate_content(
            [myfile, "\n\n", "Write a caption for this image. make sure that the caption is descriptive and perfectly describes whatever the image is trying to convey to the viewer."],
            safety_settings=[
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"}
        ])
        os.unlink(temp_file_path)
        return result.text
    
    except Exception as e:
        os.unlink(temp_file_path)
        print(f"An error occurred: {e}")
        return None
    
def limit_posts_per_month(posts, max_posts_per_month):
    """
    Limit the number of posts per month based on the user-defined maximum.
    
    :param posts: List of posts sorted by date (newest first)
    :param max_posts_per_month: Maximum number of posts allowed per month
    :return: List of posts limited to max_posts_per_month per month
    """
    limited_posts = []
    current_month = None
    month_post_count = 0
    
    for post in posts:
        post_date = datetime.strptime(post['timestamp'].split('T')[0], '%Y-%m-%d').date()
        post_month = (post_date.year, post_date.month)
        
        if post_month != current_month:
            current_month = post_month
            month_post_count = 0
        
        if month_post_count < max_posts_per_month:
            limited_posts.append(post)
            month_post_count += 1
    
    return limited_posts

def run(gemini_api_key, api_key, insta_ids, flag, max_posts, day, month, year, num_month, output_folder_path, search_hashtags=None):
    
    output_folder_path = r"{}".format(output_folder_path)
    accounts = insta_ids
    today = datetime.today().date()
    today_day = today.day
    today_month = today.month
    today_year = today.year
    day = int(today_day)
    month = int(today_month)
    year = int(today_year)
    flag = "b"
    if flag == "b":
        month_cutoff = calculate_previous_date(day, month, year, num_month)
        num_post = num_months_to_posts(num_month)
    elif flag == "a":
        month_cutoff = calculate_previous_date(day, month, year, 1000)
        num_post = max_posts
    else:
        print("Invalid flag")
        return
    list_of_account_dataframes = []
    list_of_posts_dataframes = []

    
    for account in accounts:
        
        posts_df = pd.DataFrame(columns=["Platform", "Caption", "Hashtags", "Alt Text", "Post Type", "Post/Image URL", "Likes", "Comment count", "Day", "Month", "Year", "Company Name", "Is Sponsored", "Generated Caption"])

        account_df = pd.DataFrame(columns=["User Name", "Full Name", "Biography", "Subscriber Count", "Following Count", "Reels Count - video", "Posts Count - image"])
        profile_client, profile_ws, profile_wb = setup_profile(api_key)

        row = 2 ## starting from row 2
        profile_run_input = { "usernames": [account] }
        profile_run = profile_client.actor("apify/instagram-profile-scraper").call(run_input=profile_run_input)

        for item in profile_client.dataset(profile_run["defaultDatasetId"]).iterate_items():
            
            user_name = item.get("username")
            full_name = item.get("fullName")
            bio = item.get("biography")
            followers = item.get("followersCount")
            following = item.get("followsCount")
            n_reels = item.get("igtvVideoCount")
            posts = item.get("postsCount")
            
            profile_ws.cell(row=row, column=1, value=str(user_name))
            profile_ws.cell(row=row, column=2, value=str(full_name))
            profile_ws.cell(row=row, column=3, value=str(bio))
            profile_ws.cell(row=row, column=4, value=int(followers))
            profile_ws.cell(row=row, column=5, value=int(following))
            profile_ws.cell(row=row, column=6, value=int(n_reels))
            profile_ws.cell(row=row, column=7, value=int(posts))
            row += 1
            row_data = (str(user_name), str(full_name), str(bio), int(followers), int(following), int(n_reels), int(posts), )
            account_df.loc[len(account_df)] = row_data
        excel_filename = f"instagram_{account}_profile.xlsx"
        save_path = f"{output_folder_path}/{excel_filename}"
        profile_wb.save(save_path)
                        
        row = 2
         
        client, ws, wb = setup(api_key)
        input_url = f"https://www.instagram.com/{account}/?hl=en"
        run_input = {
        "directUrls": [input_url],
        "resultsType": "posts",
        "resultsLimit": num_post,
        "searchType": "hashtag",
        "searchLimit": 1,
        }
        run = client.actor("apify/instagram-scraper").call(run_input=run_input)
        all_posts = list(client.dataset(run["defaultDatasetId"]).iterate_items())
        all_posts.sort(key=lambda x: x['timestamp'], reverse=True)
        limited_posts = limit_posts_per_month(all_posts, 3)
        
        for item in limited_posts:
            ## Post text fields

            caption = item.get("caption")
            hashtags = item.get("hashtags")
            alt = item.get("alt")

            ## Post image fields

            post_type = item.get("type")
            displayUrl = item.get("displayUrl")
            response = requests.get(displayUrl)
            img_data = BytesIO(response.content)

            ## Post engagement fields

            likesCount = item.get("likesCount")
            commentsCount = item.get("commentsCount")

            timestamp = item.get("timestamp")
            ownerFullName = item.get("ownerFullName")
            isSponsored = item.get("isSponsored")
            # url = item.get("url")

            dt = timestamp.split(".")[0]
            date_time = datetime.strptime(dt, "%Y-%m-%dT%H:%M:%S")
            date = date_time.date()
            day = int(date.day)
            month = int(date.month)
            year = int(date.year)

            if date < month_cutoff:
                break
            
            if search_hashtags is not None:
                for hashtag in search_hashtags:
                    if hashtag.lower() in hashtags.lower():
                        ## continue normal execution
                        pass
                    else:
                        continue
            
            ws.cell(row=row, column=1, value='Instagram')
            ws.cell(row=row, column=2, value=str(caption))
            ws.cell(row=row, column=3, value=str(hashtags))
            ws.cell(row=row, column=4, value=str(alt))
            ws.cell(row=row, column=5, value=str(post_type))
            ws.cell(row=row, column=6, value=str(displayUrl))
            ws.cell(row=row, column=7, value=likesCount)
            ws.cell(row=row, column=8, value=commentsCount)
            ws.cell(row=row, column=9, value=int(day))
            ws.cell(row=row, column=10, value=int(month))
            ws.cell(row=row, column=11, value=int(year))
            ws.cell(row=row, column=12, value=ownerFullName)
            ws.cell(row=row, column=13, value=str(isSponsored))

            gen_caption = get_post_text(displayUrl, gemini_api_key)
            if gen_caption:
                ws.cell(row=row, column=12, value=gen_caption)
            else:
                ws.cell(row=row, column=12, value="Unable to generate caption")

            row_data = ('Instagram', str(caption), str(hashtags), str(alt), str(post_type), str(displayUrl), int(likesCount), int(commentsCount), day, month, year, str(ownerFullName), str(isSponsored), gen_caption)
            
            posts_df.loc[len(posts_df)] = row_data
            row += 1

        list_of_posts_dataframes.append(posts_df)
        list_of_account_dataframes.append(account_df)
        excel_filename_2 = f"instagram_{account}_posts.xlsx"
        save_path = f"{output_folder_path}/{excel_filename_2}"
        wb.save(save_path)
    
    return list_of_account_dataframes, list_of_posts_dataframes, excel_filename, excel_filename_2