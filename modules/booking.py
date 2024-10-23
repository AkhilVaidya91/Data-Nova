####################### IMPORTING REQIRED LIBRARIES #######################
import numpy as np
import pandas as pd
import requests
import os
import tempfile
import google.generativeai as genai
from apify_client import ApifyClient
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import random

####################### FUNCTION DEFINITIONS #######################
def remove_comma(text):
    text =  text.replace(",", " ")
    return str(text)

def setup(api_key):
    client = ApifyClient(api_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Booking-dot-com Hotel Reviews"

    headers = ["Platform", "User Name", "User Location", "Room info", "Stay Month", "Stay Year", "Stay Length", "Review Title", "Review Rating", "Review - liked", "Review - disliked"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return client, ws, wb


def run(api_key, links, max_posts, output_folder_path):
    
    output_folder_path = r"{}".format(output_folder_path)

    for link in links:
        row = 2
        df = pd.DataFrame(columns=["Platform", "User Name", "User Location", "Room info", "Stay Month", "Stay Year", "Stay Length", "Review Title", "Review Rating", "Review - liked", "Review - disliked"])
        client, ws, wb = setup(api_key)
        run_input = {
            "startUrls": [{ "url": link }],
            "maxReviewsPerHotel": max_posts,
            "sortReviewsBy": "f_relevance",
            "reviewScores": ["ALL"],
        }
        run = client.actor("voyager/booking-reviews-scraper").call(run_input=run_input)
        
        for item in client.dataset(run["defaultDatasetId"]).iterate_items():

            user_name = item.get("userName")
            user_location = item.get("userLocation")
            user_location = remove_comma(user_location)
            room_info = item.get("roomInfo")
            room_info = remove_comma(room_info)
            stay_date = item.get("stayDate")
            stay_date = stay_date.split(" ")
            stay_month = stay_date[0]
            stay_year = stay_date[1]
            stay_length = item.get("stayLength")
            stay_length = remove_comma(stay_length)
            review_title = item.get("reviewTitle")
            review_title = remove_comma(review_title)
            rating = item.get("rating")
            review_text_parts = item.get("reviewTextParts")
            liked = review_text_parts.get("Liked")
            liked = remove_comma(liked)
            disliked = review_text_parts.get("Disliked")
            disliked = remove_comma(disliked)
            

            ws.cell(row=row, column=1, value='Booking.com')
            ws.cell(row=row, column=2, value=str(user_name))
            ws.cell(row=row, column=3, value=str(user_location))
            ws.cell(row=row, column=4, value=str(room_info))
            ws.cell(row=row, column=5, value=str(stay_month))
            ws.cell(row=row, column=6, value=int(stay_year))
            ws.cell(row=row, column=7, value=str(stay_length))
            ws.cell(row=row, column=8, value=str(review_title))
            ws.cell(row=row, column=9, value=str(rating))
            ws.cell(row=row, column=10, value=str(liked))
            ws.cell(row=row, column=12, value=str(disliked))

            row_values = ('Booking.com', str(user_name), str(user_location), str(room_info), str(stay_month), int(stay_year), str(stay_length), str(review_title), str(rating), str(liked), str(disliked))
            df.loc[len(df)] = row_values
            row += 1
        random_number = random.randint(1, 100000)
        excel_filename = f"booking_{random_number}_reviews.xlsx"
        ## filename also includes timestamp
        excel_filename = f"booking_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{random_number}_reviews.xlsx"
        save_path = f"{output_folder_path}/{excel_filename}"
        wb.save(save_path)
    return df, excel_filename