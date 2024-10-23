####################### IMPORTING REQIRED LIBRARIES #######################
import numpy as np
import pandas as pd
import requests
import os
import tempfile
import google.generativeai as genai
from apify_client import ApifyClient
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

####################### FUNCTION DEFINITIONS #######################

def remove_comma(text):
    text =  text.replace(",", " ")
    return str(text)

def setup(api_key):
    client = ApifyClient(api_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Flickr Data"
    headers = ["Platform", "Title", "Favourite Count", "Comment Count", "Media Type", "Media Content"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return client, ws, wb

def get_post_text(post_url, api_key):
    genai.configure(api_key=api_key)
    response = requests.get(post_url)
    image_data = response.content

    # Create a unique temporary file for each request
    temp_dir = tempfile.mkdtemp()
    temp_file_path = os.path.join(temp_dir, f"temp_{os.getpid()}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.jpg")

    try:
        with open(temp_file_path, 'wb') as temp_file:
            temp_file.write(image_data)
        
        myfile = genai.upload_file(temp_file_path)
        model = genai.GenerativeModel("gemini-1.5-flash")
        result = model.generate_content(
            [myfile, "\n\n", "Write a caption for this image. make sure that the caption is descriptive and perfectly describes whatever the image is trying to convey to the viewer."],
            safety_settings=[
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"}
        ])
        return result.text

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

    finally:
        # Clean up both the file and the temporary directory
        try:
            os.unlink(temp_file_path)
            os.rmdir(temp_dir)
        except OSError:
            pass


def run(apify_api_key, gemini_api_key, query, max_posts, output_folder_path):
    
    output_folder_path = r"{}".format(output_folder_path)

    for link in query:
        row = 2
        df = pd.DataFrame(columns=["Platform", "Title", "Favourite Count", "Comment Count", "Media Type", "Media Content"])
        client, ws, wb = setup(apify_api_key)
        run_input = {
            "searchUrls": [],
            "searchQueries": [link],
            "desiredMediaCount": max_posts,
            "sortType": "RELEVANCE",
            "licenseType": "ALL",
            "mediaMinSize": "SMALL",
            "dateSearchType": "DATE_TAKEN",
            "proxyConfiguration": { "useApifyProxy": True },
        }

        # Run the Actor and wait for it to finish
        run = client.actor("web.harvester/flickr-scraper").call(run_input=run_input)
        
        for item in client.dataset(run["defaultDatasetId"]).iterate_items():

            title = item.get("title")
            title = remove_comma(title)
            favs = item.get("count_faves")
            comment_count = item.get("count_comments")
            media_type = item.get("media_type")

            images = item.get("images")

            for image in images:
                image_url = image.get("url")
                image_caption = get_post_text(image_url, gemini_api_key)
            
                ["Platform", "Title", "Favourite Count", "Comment Count", "Media Type", "Media Content"]
                ws.cell(row=row, column=1, value='Flickr')
                ws.cell(row=row, column=2, value=str(title))
                ws.cell(row=row, column=3, value=str(favs))
                ws.cell(row=row, column=4, value=str(comment_count))
                ws.cell(row=row, column=5, value=str(media_type))
                ws.cell(row=row, column=6, value=str(image_url))

                row_values = ('Flickr', str(title), str(favs), str(comment_count), str(media_type), str(image_caption))
                df.loc[len(df)] = row_values
                row += 1
        excel_filename = f"flickr_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{link}_data.xlsx"
        save_path = f"{output_folder_path}/{excel_filename}"
        wb.save(save_path)
    return df, excel_filename