####################### IMPORTING REQIRED LIBRARIES #######################
import numpy as np
import pandas as pd
import requests
import os
import tempfile
import google.generativeai as genai
from apify_client import ApifyClient
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

####################### FUNCTION DEFINITIONS #######################
def setup(api_key):
    client = ApifyClient(api_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tripadvisor Hotel Reviews"

    headers = ["Platform", "Date - Year", "Date - Month", "Reviewed from Device", "Rating", "Helpful Votes", "Review Text", "User Location", "Reviewer info - number of total reviews", "Reviewer info - number of cities visited", "Reviewer info - Helpful Count", "Photos text caption", "Place", "Location", "Address", "City", "State", "Country"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return client, ws, wb


def get_post_text(post_url, api_key):
    genai.configure(api_key=api_key)
    response = requests.get(post_url)
    image_data = response.content

    # Create a unique temporary file for each request
    temp_dir = tempfile.mkdtemp()
    temp_file_path = os.path.join(temp_dir, f"temp_{os.getpid()}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.jpg")

    try:
        with open(temp_file_path, 'wb') as temp_file:
            temp_file.write(image_data)
        
        myfile = genai.upload_file(temp_file_path)
        model = genai.GenerativeModel("gemini-1.5-flash")
        result = model.generate_content(
            [myfile, "\n\n", "Write a caption for this image. make sure that the caption is descriptive and perfectly describes whatever the image is trying to convey to the viewer."],
            safety_settings=[
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"}
        ])
        return result.text

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

    finally:
        # Clean up both the file and the temporary directory
        try:
            os.unlink(temp_file_path)
            os.rmdir(temp_dir)
        except OSError:
            pass

def run(gemini_api_key, api_key, links, max_posts, output_folder_path):
    
    output_folder_path = r"{}".format(output_folder_path)
    df = pd.DataFrame(columns=["Platform", "Date - Year", "Date - Month", "Reviewed from Device", "Rating", "Helpful Votes", "Review Text", "User Location", "Reviewer info - number of total reviews", "Reviewer info - number of cities visited", "Reviewer info - Helpful Count", "Photos text caption", "Place", "Location", "Address", "City", "State", "Country"])

    for link in links:
        row = 2
         
        client, ws, wb = setup(api_key)
        run_input = {
            "startUrls": [{ "url": link }],
            "maxItemsPerQuery": max_posts,
            "reviewRatings": ["ALL_REVIEW_RATINGS"],
            "reviewsLanguages": ["ALL_REVIEW_LANGUAGES"],
        }
        run = client.actor("maxcopell/tripadvisor-reviews").call(run_input=run_input)
        
        for item in client.dataset(run["defaultDatasetId"]).iterate_items():

            published_date = item.get("publishedDate")
            published_date = published_date.split("-")
            year = published_date[0]
            month = published_date[1]
            platform = item.get("publishedPlatform")
            rating = item.get("rating")
            helpful_votes = item.get("helpfulVotes")
            text = item.get("text")

            user = item.get("user")
            contributions = user.get("contributions")
            loc = user.get("userLocation")
            if loc is not None:
                loc_name = loc.get("name")
            else:
                loc_name = 'No user location'

            reviews = contributions.get("reviews")
            cities = contributions.get("reviewCityCount")
            helpful_contributions = contributions.get("helpfulVotes")

            photos = item.get("photos")
            photos_links = []
            for photo in photos:
                phto_link = photo.get("image")
                photos_links.append(phto_link)
            
            place = item.get("placeInfo")
            place_name = place.get("name")
            place_location = place.get("locationString")
            place_address = place.get("address")
            address = place.get("addressObj")

            city = address.get("city")
            state = address.get("state")
            country = address.get("country")

            

            ws.cell(row=row, column=1, value='TripAdvisor')
            ws.cell(row=row, column=2, value=int(year))
            ws.cell(row=row, column=3, value=int(month))
            ws.cell(row=row, column=4, value=str(platform))
            ws.cell(row=row, column=5, value=str(rating))
            ws.cell(row=row, column=6, value=str(helpful_votes))
            ws.cell(row=row, column=7, value=str(text))
            ws.cell(row=row, column=8, value=str(loc_name))
            ws.cell(row=row, column=9, value=str(reviews))
            ws.cell(row=row, column=10, value=str(cities))
            ws.cell(row=row, column=11, value=str(helpful_contributions))
            ws.cell(row=row, column=13, value=str(place_name))
            ws.cell(row=row, column=14, value=str(place_location))
            ws.cell(row=row, column=15, value=str(place_address))
            ws.cell(row=row, column=16, value=str(city))
            ws.cell(row=row, column=17, value=str(state))
            ws.cell(row=row, column=18, value=str(country))
            captions = []
            for i, link in enumerate(photos_links):
                image_caption = get_post_text(link, gemini_api_key)
                if image_caption:
                    text = f"Image {i+1}: {image_caption}"
                else:
                    pass
                captions.append(text)

            ws.cell(row=row, column=12, value=str(captions))
            row_data = ("TripAdvisor", int(year), int(month), str(platform), str(rating), str(helpful_votes), str(text), str(loc_name), str(reviews), str(cities), str(helpful_contributions), str(captions), str(place_name), str(place_location), str(place_address), str(city), str(state), str(country))
            df.loc[row] = row_data
            row += 1
        excel_filename = f"tripadvisor_{place_name}_reviews.xlsx"
        save_path = f"{output_folder_path}/{excel_filename}"
        wb.save(save_path)
    return df, excel_filename
