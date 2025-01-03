import numpy as np
import pandas as pd
from apify_client import ApifyClient
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
import requests
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from pymongo import MongoClient
import os

MONGO_URI = os.getenv('MONGO_URI')
MONGO_URI = "mongodb+srv://akhilvaidya22:qN2dxc1cpwD64TeI@digital-nova.cbbsn.mongodb.net/?retryWrites=true&w=majority&appName=digital-nova"
client = MongoClient(MONGO_URI)
db = client['digital_nova']
output_files_collection = db['output_files']

####################### FUNCTION DEFINITIONS #######################

def setup(api_key):
    client = ApifyClient(api_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon Reviews"
    headers = ["Review source", "Rating", "Review Title", "Review Reaction", "Country", "Date", "Month", "Year", "Description", "Product Variant"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return client, ws, wb

def get_demographics(string):
    if len(string) > 10:
        strings = string.split(" ")
        country = strings[2]
        date = strings[4]
        month = strings[5]
        year = strings[6]

        return str(country), int(date), str(month), int(year)
    else:
        return None, None, None, None

def run(api_key, product_urls, output_folder_path, username, max_reviews=100):
    
    df = pd.DataFrame(columns=["Review source", "Rating", "Review Title", "Review Reaction", "Country", "Date", "Month", "Year", "Description", "Product Variant"])
    output_folder_path = r"{}".format(output_folder_path)
    products = product_urls
    
    num = 1
    for product in products:
        row = 2
         
        client, ws, wb = setup(api_key)
        input_url = product
        run_input = {
            "productUrls": [{ "url": input_url }],
            "maxReviews": max_reviews,
            "filterByRatings": ["allStars"],
            "proxyCountry": "AUTO_SELECT_PROXY_COUNTRY",
        }
        run = client.actor("junglee/amazon-reviews-scraper").call(run_input=run_input)
        
        for item in client.dataset(run["defaultDatasetId"]).iterate_items():
            

            rating = item.get("ratingScore")
            review_title = item.get("reviewTitle")
            review_reaction = item.get("reviewReaction")
            demographics = item.get("reviewedIn")
            country, date, month, year = get_demographics(demographics)
            desc = item.get("reviewDescription")
            product_variant = item.get("variant")

            ws.cell(row=row, column=1, value='Amazon')
            ws.cell(row=row, column=2, value=str(rating))
            ws.cell(row=row, column=3, value=str(review_title))
            ws.cell(row=row, column=4, value=str(review_reaction))
            ws.cell(row=row, column=5, value=str(country))
            ws.cell(row=row, column=6, value=int(date))
            ws.cell(row=row, column=7, value=str(month))
            ws.cell(row=row, column=8, value=int(year))
            ws.cell(row=row, column=9, value=str(desc))
            ws.cell(row=row, column=10, value=str(product_variant))
            
            row_data = ("Amazon", str(rating), str(review_title), str(review_reaction), str(country), int(date), str(month), int(year), str(desc), str(product_variant))
            df.loc[len(df)] = row_data
            row += 1
        
        excel_filename = f"amazon_{num}_reviews.xlsx"
        ## filename also includes timestamp
        excel_filename = f"{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}_{excel_filename}"
        # save_path = f"{output_folder_path}/{excel_filename}"
        # wb.save(save_path)
        num = num + 1

        output_files_collection.insert_one({
            'username': username,
            'file_type': 'amazon_reviews',
            'filename': excel_filename,
            # 'path': save_path,
            'created_at': datetime.now()
        })
    return df, excel_filename