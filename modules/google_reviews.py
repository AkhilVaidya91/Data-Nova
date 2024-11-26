## google maps reviews scraping using - compass/Google-Maps-Reviews-Scraper

####################### IMPORTING REQIRED LIBRARIES #######################
import numpy as np
import pandas as pd
import requests
import os
import tempfile
import google.generativeai as genai
from apify_client import ApifyClient
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import random
import streamlit as st
from pymongo import MongoClient


MONGO_URI = os.getenv('MONGO_URI')
MONGO_URI = "mongodb+srv://akhilvaidya22:qN2dxc1cpwD64TeI@digital-nova.cbbsn.mongodb.net/?retryWrites=true&w=majority&appName=digital-nova"
client = MongoClient(MONGO_URI)
db = client['digital_nova']
output_files_collection = db['output_files']

####################### FUNCTION DEFINITIONS #######################
def y_m_d_to_date(year, month, day):
    return datetime(year, month, day).date()

def remove_comma(text):
    text =  text.replace(",", " ")
    return str(text)

def get_post_text(post_url, api_key):
    genai.configure(api_key=api_key)
    response = requests.get(post_url)
    image_data = response.content

    # Create a unique temporary file for each request
    temp_dir = tempfile.mkdtemp()
    temp_file_path = os.path.join(temp_dir, f"temp_{os.getpid()}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.jpg")

    try:
        with open(temp_file_path, 'wb') as temp_file:
            temp_file.write(image_data)
        
        myfile = genai.upload_file(temp_file_path)
        model = genai.GenerativeModel("gemini-1.5-flash")
        result = model.generate_content(
            [myfile, "\n\n", "Write a caption for this image. make sure that the caption is descriptive and perfectly describes whatever the image is trying to convey to the viewer."],
            safety_settings=[
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"}
        ])
        return result.text

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

    finally:
        # Clean up both the file and the temporary directory
        try:
            os.unlink(temp_file_path)
            os.rmdir(temp_dir)
        except OSError:
            pass

def setup(api_key):
    client = ApifyClient(api_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Google News Articles"

    headers = ["Platform", "Title", "Link", "Published Source", "Year", "Month", "Date", "Image Caption", "Content"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return client, ws, wb

def extract_article_content(url, perplexity_api_key, max_tokens=500):

    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {perplexity_api_key}",
        "Content-Type": "application/json"
    }
    
    api_url = "https://api.perplexity.ai/chat/completions"
    
    payload = {
        "model": "llama-3.1-sonar-small-128k-online",
        "messages": [
            {
                "role": "system",
                "content": "You are a helpful assistant that extracts and summarizes news article content."
            },
            {
                "role": "user",
                "content": f"Visit this news article URL: {url} and extract its main content. Provide only the article's text content, excluding any advertisements, navigation elements, or formatting. If you cannot access the URL, please indicate that clearly. Return only a simple string of text rather than a markdown"
            }
        ],
        "temperature": 0.1,
        "max_tokens": max_tokens,
        "stream": False
    }
    
    try:
        response = requests.post(api_url, json=payload, headers=headers)
        
        if response.status_code != 200:
            print(f"Error {response.status_code}: {response.text}")
            return None
            
        content = response.json()
        
        if 'choices' in content and len(content['choices']) > 0:
            extracted_text = content['choices'][0]['message']['content']
            return extracted_text
        else:
            print(f"Unexpected response format: {content}")
            return None
            
    except requests.exceptions.RequestException as e:
        print(f"Request error for {url}: {str(e)}")
        return None
    except Exception as e:
        print(f"Unexpected error processing {url}: {str(e)}")
        return None


def run(api_key, gemini_api_key, query, max_reviews, start_date, output_folder_path, username):
    
    output_folder_path = r"{}".format(output_folder_path)

    row = 2
    df = pd.DataFrame(columns=["Platform", "Location Name", "Review Link", "Stars", "Review Text", "Likes Count", "Date", "Month", "Year", "Image Caption", "Number of Reviewer Reviews"])
    client, ws, wb = setup(api_key)
    run_input = {
        "startUrls": [{"url": query}],
        "maxReviews": max_reviews,
        "language": "en",
        "reviewsSort": "newest",
        "reviewsStartDate": start_date,
    }

    # Run the Actor and wait for it to finish
    run = client.actor("compass/google-maps-reviews-scraper").call(run_input=run_input)
    
    for item in client.dataset(run["defaultDatasetId"]).iterate_items():
        
        location_name = item.get("title")
        reviewer_name = item.get("name")
        review_text = item.get("text")
        if review_text is None:
            review_text = "No review text available"

        published_date = item.get("publishedAtDate")
        if published_date:
            date = published_date.split("T")[0]
            date = date.split("-")

            published_year = date[0]
            published_year = int(published_year)
            published_month = date[1]
            published_month = int(published_month)
            published_day = date[2]
            published_day = int(published_day)
        else:
            published_year = 0
            published_month = 0
            published_day = 0

        likes_count = item.get("likesCount")
        if likes_count:
            likes_count = int(likes_count)
        else:
            likes_count = 0
        review_link = item.get("reviewUrl")
        if review_link is None:
            review_link = "No link available"
        number_of_reviewer_reviews = item.get("reviewerNumberOfReviews")
        if number_of_reviewer_reviews:
            number_of_reviewer_reviews = int(number_of_reviewer_reviews)
        else:
            number_of_reviewer_reviews = 0
        stars = item.get("stars")
        if stars:
            stars = int(stars)
        else:
            stars = 0

        review_image_urls = item.get("reviewImageUrls")
        image_captions = ""
        if review_image_urls:
            for image_url in review_image_urls:
                image_caption = get_post_text(image_url, gemini_api_key)
                if image_caption:
                    image_caption = remove_comma(image_caption)
                    image_captions += image_caption + " | "
                else:
                    image_caption = "No caption available"
                    image_captions += image_caption + " | "
        else:
            image_captions = "No image available"
        
        ws.cell(row=row, column=1, value= 'Google Reviews')
        ws.cell(row=row, column=2, value= location_name)
        ws.cell(row=row, column=3, value= str(review_link))
        ws.cell(row=row, column=4, value= stars)
        ws.cell(row=row, column=5, value= review_text)
        ws.cell(row=row, column=6, value= likes_count)
        ws.cell(row=row, column=7, value= published_date)
        ws.cell(row=row, column=8, value= published_month)
        ws.cell(row=row, column=9, value= published_year)
        ws.cell(row=row, column=10, value= image_captions)
        ws.cell(row=row, column=11, value= number_of_reviewer_reviews)

        row_values = ("Google Reviews", location_name, review_link, stars, review_text, likes_count, published_day, published_month, published_year, image_captions, number_of_reviewer_reviews)
        df.loc[len(df)] = row_values
        row += 1
    excel_filename = f"google_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{location_name}_reviews.xlsx"

    output_files_collection.insert_one({
    'username': username,
    'file_type': 'Google Reviews',
    'file_name': excel_filename,
    'timestamp': datetime.now()
    })
    return df, excel_filename