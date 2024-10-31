####################### IMPORTING REQIRED LIBRARIES #######################
import numpy as np
import pandas as pd
import requests
import os
import tempfile
import google.generativeai as genai
from apify_client import ApifyClient
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import random
import streamlit as st
from pymongo import MongoClient


MONGO_URI = os.getenv('MONGO_URI')
client = MongoClient(MONGO_URI)
db = client['digital_nova']
output_files_collection = db['output_files']

####################### FUNCTION DEFINITIONS #######################
def y_m_d_to_date(year, month, day):
    return datetime(year, month, day).date()

def remove_comma(text):
    text =  text.replace(",", " ")
    return str(text)

def get_post_text(post_url, api_key):
    genai.configure(api_key=api_key)
    response = requests.get(post_url)
    image_data = response.content

    # Create a unique temporary file for each request
    temp_dir = tempfile.mkdtemp()
    temp_file_path = os.path.join(temp_dir, f"temp_{os.getpid()}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.jpg")

    try:
        with open(temp_file_path, 'wb') as temp_file:
            temp_file.write(image_data)
        
        myfile = genai.upload_file(temp_file_path)
        model = genai.GenerativeModel("gemini-1.5-flash")
        result = model.generate_content(
            [myfile, "\n\n", "Write a caption for this image. make sure that the caption is descriptive and perfectly describes whatever the image is trying to convey to the viewer."],
            safety_settings=[
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"}
        ])
        return result.text

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

    finally:
        # Clean up both the file and the temporary directory
        try:
            os.unlink(temp_file_path)
            os.rmdir(temp_dir)
        except OSError:
            pass

def setup(api_key):
    client = ApifyClient(api_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Google News Articles"

    headers = ["Platform", "Title", "Link", "Published Source", "Year", "Month", "Date", "Image Caption", "Content"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return client, ws, wb

def extract_article_content(url, perplexity_api_key, max_tokens=500):

    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {perplexity_api_key}",
        "Content-Type": "application/json"
    }
    
    api_url = "https://api.perplexity.ai/chat/completions"
    
    payload = {
        "model": "llama-3.1-sonar-small-128k-online",
        "messages": [
            {
                "role": "system",
                "content": "You are a helpful assistant that extracts and summarizes news article content."
            },
            {
                "role": "user",
                "content": f"Visit this news article URL: {url} and extract its main content. Provide only the article's text content, excluding any advertisements, navigation elements, or formatting. If you cannot access the URL, please indicate that clearly. Return only a simple string of text rather than a markdown"
            }
        ],
        "temperature": 0.1,
        "max_tokens": max_tokens,
        "stream": False
    }
    
    try:
        response = requests.post(api_url, json=payload, headers=headers)
        
        if response.status_code != 200:
            print(f"Error {response.status_code}: {response.text}")
            return None
            
        content = response.json()
        
        if 'choices' in content and len(content['choices']) > 0:
            extracted_text = content['choices'][0]['message']['content']
            return extracted_text
        else:
            print(f"Unexpected response format: {content}")
            return None
            
    except requests.exceptions.RequestException as e:
        print(f"Request error for {url}: {str(e)}")
        return None
    except Exception as e:
        print(f"Unexpected error processing {url}: {str(e)}")
        return None


def run(api_key, gemini_api_key, perplexity_api_key, query, max_articles, start_date, end_date, output_folder_path, username):
    
    output_folder_path = r"{}".format(output_folder_path)

    row = 2
    df = pd.DataFrame(columns=["Platform", "Title", "Link", "Published Source", "Year", "Month", "Date", "Image Caption", "Content"])
    client, ws, wb = setup(api_key)
    run_input = {
        "query": str(query),
        "language": "US:en",
        "maxItems": max_articles,
        "extractImages": True,
        "dateFrom": str(start_date),
        "dateTo": str(end_date),
        "proxyConfiguration": { "useApifyProxy": True },
    }

    # Run the Actor and wait for it to finish
    run = client.actor("lhotanova/google-news-scraper").call(run_input=run_input)
    
    for item in client.dataset(run["defaultDatasetId"]).iterate_items():

        title = item.get("title")
        if title:
            title = remove_comma(title) #
        link = item.get("link") #
        source = item.get("source")
        if source:
            source = remove_comma(source) #
        publish_date = item.get("publishedAt")
        publish_date = publish_date.split("-")
        published_year = publish_date[0]    #
        published_month = publish_date[1]   #
        published_day = publish_date[2]
        published_day = published_day.split("T")[0] #
        image_link = item.get("image")
        if image_link:
            image_caption = get_post_text(image_link, gemini_api_key)   #
            if image_caption:
                image_caption = remove_comma(image_caption)
            else:
                image_caption = "No caption available"
        else:
            image_caption = "No image available"


        if link:
            description = extract_article_content(link, perplexity_api_key)
            if description:
                description = remove_comma(description) #
            else:
                description = "No content available"
        else:
            description = "No content available"
        

        ws.cell(row=row, column=1, value='Google News')
        ws.cell(row=row, column=2, value=str(title))
        ws.cell(row=row, column=3, value=str(link))
        ws.cell(row=row, column=4, value=str(source))
        ws.cell(row=row, column=5, value=int(published_year))
        ws.cell(row=row, column=6, value=int(published_month))
        ws.cell(row=row, column=7, value=int(published_day))
        ws.cell(row=row, column=8, value=str(image_caption))
        ws.cell(row=row, column=9, value=str(description))

        row_values = ('Google News', str(title), str(link), str(source), int(published_year), int(published_month), int(published_day), str(image_caption), str(description))
        df.loc[len(df)] = row_values
        row += 1
    excel_filename = f"google_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{query}_news.xlsx"
    save_path = f"{output_folder_path}/{excel_filename}"
    wb.save(save_path)

    output_files_collection.insert_one({
    'username': username,
    'file_type': 'Instagram',
    'file_name': excel_filename,
    'file_path': save_path,
    'timestamp': datetime.now()
    })
    return df, excel_filename