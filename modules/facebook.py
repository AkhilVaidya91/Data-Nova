####################### IMPORTING REQIRED LIBRARIES #######################
import numpy as np
import pandas as pd
import requests
import os
import tempfile
import google.generativeai as genai
from apify_client import ApifyClient
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from pymongo import MongoClient

MONGO_URI = os.getenv('MONGO_URI')
MONGO_URI = "mongodb+srv://akhilvaidya22:qN2dxc1cpwD64TeI@digital-nova.cbbsn.mongodb.net/?retryWrites=true&w=majority&appName=digital-nova"
client = MongoClient(MONGO_URI)
db = client['digital_nova']
output_files_collection = db['output_files']

####################### FUNCTION DEFINITIONS #######################

def remove_comma(text):
    text =  text.replace(",", " ")
    return str(text)

def setup(api_key):
    client = ApifyClient(api_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Facebook Posts"
    headers = ["Platform", "Page Name", "Date", "Month", "Year", "Time", "Likes Count", "Comments Count", "Shares Count", "Post Text", "Image Caption"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return client, ws, wb

def setup_profile(api_key):
    client = ApifyClient(api_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Facebook Profile"
    headers = ["Categories", "Info", "Likes", "Num Posts", "Title", "Address", "Page Name", "Page URL", "Contact Number"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return client, ws, wb


def get_post_text(post_url, api_key):
    if post_url is None:
        return None
    if api_key is None:
        return None
    genai.configure(api_key=api_key)
    response = requests.get(post_url)
    image_data = response.content

    # Create a unique temporary file for each request
    temp_dir = tempfile.mkdtemp()
    temp_file_path = os.path.join(temp_dir, f"temp_{os.getpid()}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.jpg")

    try:
        with open(temp_file_path, 'wb') as temp_file:
            temp_file.write(image_data)
        
        myfile = genai.upload_file(temp_file_path)
        model = genai.GenerativeModel("gemini-1.5-flash")
        result = model.generate_content(
            [myfile, "\n\n", "Write a caption for this image. make sure that the caption is descriptive and perfectly describes whatever the image is trying to convey to the viewer."],
            safety_settings=[
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"}
        ])
        return result.text

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

    finally:
        # Clean up both the file and the temporary directory
        try:
            os.unlink(temp_file_path)
            os.rmdir(temp_dir)
        except OSError:
            pass


def run(gemini_api_key, api_key, facebook_url, start_date, end_date, max_posts, output_folder_path, username):
    
    output_folder_path = r"{}".format(output_folder_path)
        
    posts_df = pd.DataFrame(columns=["Platform", "Page Name", "Date", "Month", "Year", "Time", "Likes Count", "Comments Count", "Shares Count", "Post Text", "Image Caption"])

    account_df = pd.DataFrame(columns=["Categories", "Info", "Likes", "Num Posts", "Title", "Address", "Page Name", "Page URL", "Contact Number"])
    profile_client, profile_ws, profile_wb = setup_profile(api_key)

    row = 2 ## starting from row 2
    profile_run_input = { "startUrls": [{ "url": facebook_url}] }
    profile_run = profile_client.actor("apify/facebook-pages-scraper").call(run_input=profile_run_input)

    for item in profile_client.dataset(profile_run["defaultDatasetId"]).iterate_items():
        
        categories = item.get("categories")
        categories_str = ""
        for category in categories:
            categories_str += category + ", "
        if len(categories_str) > 0:
            categories_str = categories_str[:-2]

        info = item.get("info")
        info_str = ""
        for inf in info:
            info_str += inf + ", "
        if len(info_str) > 0:
            info_str = info_str[:-2]

        likes = item.get("likes")
        posts = item.get("posts")
        if posts == None:
            num_posts = 0
        else:
            num_posts = len(posts)

        title = item.get("title")
        address = item.get("address")
        pageName = item.get("pageName")
        page_url = item.get("pageUrl")
        phone = item.get("phone")
        
        profile_ws.cell(row=row, column=1, value=str(categories_str))
        profile_ws.cell(row=row, column=2, value=str(info_str))
        profile_ws.cell(row=row, column=3, value=int(likes))
        profile_ws.cell(row=row, column=4, value=int(num_posts))
        profile_ws.cell(row=row, column=5, value=str(title))
        profile_ws.cell(row=row, column=6, value=str(address))
        profile_ws.cell(row=row, column=7, value=str(pageName))
        profile_ws.cell(row=row, column=8, value=str(page_url))
        profile_ws.cell(row=row, column=9, value=str(phone))
        row += 1
        row_data = (str(categories_str), str(info_str), int(likes), int(num_posts), str(title), str(address), str(pageName), str(page_url), str(phone))
        account_df.loc[len(account_df)] = row_data
    excel_filename = f"facebook_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{pageName}_profile.xlsx"
    # save_path = f"{output_folder_path}/{excel_filename}"
    # profile_wb.save(save_path)

    output_files_collection.insert_one({
        'username': username,
        'file_type': 'Facebook-Profile',
        'file_name': excel_filename,
        # 'file_path': save_path,
        'timestamp': datetime.now()
    })
                    
    row = 2
        
    client, ws, wb = setup(api_key)
    run_input = {
        "startUrls": [{ "url": facebook_url }],
        "resultsLimit": max_posts,
        'onlyPostsNewerThan': start_date,
        'onlyPostsOlderThan': end_date
    }
    run = client.actor("apify/facebook-posts-scraper").call(run_input=run_input)
    
    for item in client.dataset(run["defaultDatasetId"]).iterate_items():
        ## Post text fields

        pageName = item.get("pageName")
        time = item.get("time")
        time = time.split("T")

        date = time[0]
        date = date.split("-")
        year = date[0]
        month = date[1]
        date = date[2]
        time = time[1]
        time = time.split(".")
        time = time[0]

        likes = item.get("likes")

        comments = item.get("comments")
        shares = item.get("shares")


        text = item.get("text")
        media = item.get("media")
        caption = ""
        if media is not None:
            for med in media:
                thumb = med.get("thumbnail")
                
                if thumb == None:
                    continue
                thumb_caption = get_post_text(thumb, gemini_api_key)
                if thumb_caption != None:
                    
                    caption = caption + thumb_caption

        image_caption = caption


        ws.cell(row=row, column=1, value='Facebook')
        ws.cell(row=row, column=2, value=str(pageName))
        # ws.cell(row=row, column=3, value=str(day))
        ws.cell(row=row, column=4, value=str(date))
        ws.cell(row=row, column=5, value=str(month))
        ws.cell(row=row, column=6, value=int(year))
        ws.cell(row=row, column=7, value=str(time))
        ws.cell(row=row, column=8, value=str(likes))
        ws.cell(row=row, column=9, value=str(comments))
        ws.cell(row=row, column=10, value=str(shares))
        ws.cell(row=row, column=11, value=str(text))
        ws.cell(row=row, column=12, value=str(image_caption))


        row_data = ('Facebook', str(pageName), str(date), str(month), int(year), str(time), str(likes), str(comments), str(shares), str(text), str(image_caption))
        
        posts_df.loc[len(posts_df)] = row_data
        row += 1
    
    excel_filename_2 = f"facebook_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{pageName}_posts.xlsx"
    # save_path = f"{output_folder_path}/{excel_filename_2}"
    # wb.save(save_path)

    output_files_collection.insert_one({
        'username': username,
        'file_type': 'Facebook-Poosts',
        'file_name': excel_filename_2,
        # 'file_path': save_path,
        'timestamp': datetime.now()
    })
    
    return account_df, posts_df, excel_filename, excel_filename_2